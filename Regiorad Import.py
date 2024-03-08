import requests
import os
import pandas as pd
from datetime import datetime
import time
import openpyxl

API_ENDPOINT = "https://api.mobidata-bw.de/sharing/gbfs/regiorad_stuttgart/station_status"
STATION_IDS = [
    "CAB:Station:958219cd-8009-45b6-8a04-1cc4ae763307",
    "CAB:Station:be796eb3-49af-494c-b3f0-dc1f0132eb7d",
    "CAB:Station:36fe6a30-d53e-4aa8-863d-e21e80fd1d0b",
    "CAB:Station:2263b695-db35-4bbf-b70a-032920fadf3f",
    "CAB:Station:59830a25-c7c6-4c9b-af21-0b67c836ba8f",
    "CAB:Station:714932fd-4836-4a2a-9a19-07a6d694274c",
    "CAB:Station:e73ff873-8f43-41b4-868a-2d3e00f2cfbb",
    "CAB:Station:c070e40c-98f6-45d6-9543-90ceceef63af",
    "CAB:Station:ca41b63e-9046-442f-880e-d0e9186e507a",
    "CAB:Station:e375c577-ca9b-4f3d-b052-582f4eeeb0a8",
    "CAB:Station:7046d131-47fd-40bd-ba63-f52590d2ee8f",
    "CAB:Station:69a20737-5c07-4ab4-ad8a-4fcba602ef6b",
    "CAB:Station:b207ab7a-f23d-40f8-aa26-5d02077fc6f9",
    "CAB:Station:f384359e-349d-47cf-aec3-031f20b8e0a0"
]

excel_file_path = 'station_data_analysis.xlsx'

def fetch_and_filter_data():
    response = requests.get(API_ENDPOINT)
    if response.status_code == 200:
        data = response.json()
        filtered_data = [station for station in data['data']['stations'] if station['station_id'] in STATION_IDS]
        return filtered_data
    else:
        print(f"Fehler bei der API-Anfrage. Statuscode: {response.status_code}")
        return []

def save_data_to_csv(data):
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    csv_file_path = f'station_data_{timestamp}.csv'
    dataframe = pd.DataFrame(data)
    dataframe = dataframe[['station_id', 'num_bikes_available', 'last_reported']]
    dataframe['timestamp'] = datetime.now()
    dataframe.to_csv(csv_file_path, index=False)
    print(f"Daten in {csv_file_path} gespeichert.")
    return csv_file_path


def append_csv_to_excel(csv_file_path):
    data = pd.read_csv(csv_file_path)
    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        if writer.sheets['Station Data'].max_row > 1:
            data.to_excel(writer, index=False, sheet_name='Station Data', startrow=writer.sheets['Station Data'].max_row, header=False)
        else:
            data.to_excel(writer, index=False, sheet_name='Station Data', startrow=writer.sheets['Station Data'].max_row)
    os.remove(csv_file_path)
    print(f"Daten aus {csv_file_path} in {excel_file_path} übertragen und CSV gelöscht.")


def main():
    while True:
        filtered_data = fetch_and_filter_data()
        if filtered_data:
            csv_file_path = save_data_to_csv(filtered_data)
            append_csv_to_excel(csv_file_path)
        else:
            print("Keine Daten für die vorgegebenen Station-IDs gefunden.")
        print("Warte 15 min bis zur nächsten Abfrage.")
        time.sleep(900)

if __name__ == "__main__":
    main()